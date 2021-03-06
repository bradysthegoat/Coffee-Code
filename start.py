# -*- coding: utf-8 -*-
"""
Created on Tue Aug 21 11:29:20 2018

This file contains the main code for setting up the data analysis. The main loop aligns the 3 excel files in order to
send the only the required data for analysis.

ver 0.01

@author: PKellicker
"""

import pandas as pd
import numpy as np 
import os
import re
import argparse 
from filter import find_curr_recipe
from parse_data import parse_data
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



# searches for the correct line to start so headers are correct. Used when number of lines to skip is dynamic
def skip_to(fle, line, folder, **kwargs):
    os.chdir(folder)
    if os.stat(fle).st_size == 0:
        raise ValueError("File is empty")
    with open(fle) as f:
        pos = 0
        cur_line = f.readline()
        while not cur_line.startswith(line):
            pos = f.tell()
            cur_line = f.readline()
        f.seek(pos)
        return pd.read_csv(f, **kwargs)
    

    
    
    
# ********************************************************************* #
def run(args):
    
    ### Setup ###
    # define outside cwd dir
    base = os.getcwd() #+ "\\CDT Test"
    csv_folder = base + "\\" + args.data #"+\\Use 1" #
    
    #csv_folder = "C:\\Users\\pkellicker\\Desktop\\Coffee Code\\Git Code\\Excel\\Use 2" # C:\Users\pkellicker\Desktop\Coffee Code\Code
   # base = "C:\\Users\\pkellicker\\Desktop\\Coffee Code\\Git Code\\Excel"


    os.chdir(base)

    # import static files
    cesar_name = args.kpi #= "CP300 & CF90 Cup temp-volume.xlsm" # 
    recipe_name = args.recipe #= "CP300 EB2.xlsx" #

    cesar = pd.read_excel(cesar_name)
    recipe = pd.read_excel(recipe_name, skiprows = 1)

    # ********************************************************************* #  

    # * * * work cesar to correspond w/ csv * * * #
    cesar_csv = cesar.filter(items = ['CSV Filename']) # Create df of cesar with csv filename only *** does skip Tea line
    cesar_csv = cesar_csv.dropna() # Clean up df
    
    # find length based on this column - should be more accurate
    cesar_length = cesar.filter(items = ["Pre Test Chamber Weight grams"])
    cesar_length = cesar_length.dropna()
    file_count = cesar_length.shape[0] # get number of csv files
    
    names = cesar_csv.values # I want a string array of the names so I can feed them into the fn skip_to

    # ********************************************************************* #

    # * * * work recipe to correspond w/ csv * * * #
    tea = 0 # flag for when we hit tea

    # create array of dates -> field in which Tea brew start is noted
    cesar_date = cesar.filter(items = ['Date'])
    cesar_date = cesar_date.dropna()
    dates = cesar_date.values
    
    # find number of date entries to avoid index error
    num_dates = cesar_date.shape[0]
    


    # ********************************************************************* #

    count = 0
    array = np.zeros((file_count, 6))
    end_time = args.end

    # Load workbook for storing data & find correct columns' index
    wb = load_workbook(base + "\\" + cesar_name, read_only = False, keep_vba = True)
    ws = wb['Data']


    # find column number
    avg_bot_col = cesar.columns.get_loc("Avg Bot") + 1
    avg_mid_col = avg_bot_col -1
    avg_top_col = avg_bot_col -2 
    time_zero_col = cesar.columns.get_loc("Time Zero Temp") + 1
    
    #convert column number to letter
    avg_bot_col_let = get_column_letter(avg_bot_col)
    avg_mid_col_let = get_column_letter(avg_mid_col)
    avg_top_col_let = get_column_letter(avg_top_col)
   #time_zero_col_let = get_column_letter(time_zero_col)
    

    ### Loop ###
    for i in range(0,file_count):
   
        # check for when tea starts - also for incrementing since one line in xlsx is for  *** careful it needs to find tea i.e. start b4 tea line
        if i < num_dates:
            if tea == 0:
                if dates[i] == "EB2 TEA BREWS" :
                    tea = 1
    
        file_temp = names[i] + ".csv" # add the .csv extension to file name
        file_temp = str(file_temp)[2:-2] # remove first 2 and last 2 chars -> ['']
    
        if os.path.isfile(csv_folder + "\\" + file_temp) == False: # check to see if file exists
            continue
        
        data_curr = skip_to(file_temp, "Number", csv_folder) # current csv file working with
    
        data_table = skip_to(file_temp, "CH", csv_folder)
        #data_f = data_table.iloc[0:10, 0:2] # get top of csv file -> channel table
        #data_table.head(8).to_excel("csv_Headers.xlsx")
    
        # create filter for current row, create dataframe of current row
        cesar_rowi = cesar['Test Number'] == i + 1 + tea
        cesar_curr = cesar[cesar_rowi] # cesar_curr = row corresponding to current csv file *** need to step by 1 when -> tea
        
        recipe_curr_1 = find_curr_recipe(cesar_curr, recipe, tea)
        recipe_curr = recipe_curr_1[0]
        brew_type = recipe_curr_1[1]
        
        data = parse_data(recipe_cur = recipe_curr, data_cur = data_curr, data_f = data_table, cesar_cur = cesar_curr, end_time = end_time, top = args.top, mid = args.mid, bot = args.bot) 
    

        array[i][0] = data[0] # bot
        array[i][1] = data[1] # mid
        array[i][2] = data[2] # top
        
        array[i][3] = data[3] # one
        array[i][4] = data[4] # two
        
    
        #store data
        #os.chdir("C:\\Users\\pkellicker\\Desktop\\Coffee Code\\Git Code\\Excel")
    
        ws.cell(row = i+2, column = avg_bot_col).value = array[i][0]
        ws.cell(row = i+2, column = avg_mid_col).value = array[i][1]
        ws.cell(row = i+2, column = avg_top_col).value = array[i][2]
        #ws.cell(row = i+2, column = avg_top_col-2).value = array[i][4]
        #ws.cell(row = i+2, column = avg_top_col-1).value = array[i][5]
        
        # input average formula into column "Time Zero Temp"
        string_1 = "=ROUND(AVERAGE(" + avg_bot_col_let + str(i+2) + "," + avg_mid_col_let + str(i+2) + "," + avg_top_col_let + str(i+2) + "),1)" # ********************changed could be bug
        ws.cell(row = i+2, column = time_zero_col).value = string_1
    
    wb.save(base + "\\" + cesar_name) 

def main():
    parser = argparse.ArgumentParser(description = "Analyze coffee test data. Need KPI, recipe, and csv files. Recipe and KPI files, and Data folder should be placed in same location as .exe. Be sure to use " " when specifying names")
    required = parser.add_argument_group("Required Arguments")
    required.add_argument("-kpi", help = "enter full name of KPI file", type = str, required = True)
    required.add_argument("-recipe", help = "enter full name of recipe file", type = str, required = True)
    required.add_argument("-data", help = "enter full name of folder containing the .csv files", type = str, required = True)
    
    optional = parser.add_argument_group("Other Optional Arguments")
    optional.add_argument("-v",  action = 'store_false', help="Run analysis based on voltage pattern. Default is to analyze by end_time", dest = "end", default = True) 
    optional.add_argument("-top", help="enter name of top channel title", type = str, default =None)
    optional.add_argument("-mid", help="enter name of mid channel title", type = str, default =None)
    optional.add_argument("-bot", help="enter name of bot channel title", type = str, default =None)
    parser.set_defaults(func=run)
    args = parser.parse_args()
    args.func(args)

 

if __name__ == "__main__":
    main()




#  * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * *  #

# tried to find Tea with regex

#cesar_date.to_excel("KPI Dates1.xlsx")
#fh = open("KPI Dates1.xlsx").read()
#match = re.search("EB2 TEA BREWS", fh)



'''
    #cols = ['CH','Signal name']
    #data = data_table[cols]
    data_f = data_table.iloc[0:10, 0:2]
    
    data = data_table.iloc[0:10, 1:2]
    data_0 = data.values
    data_1 = str(data_0)
    data_2 = data_1.split(" ")
    data_3 = pd.Series(data_2)
    
    data_filter = data_3.str.contains("BOT")
    data_bot = data_f[data_filter]
    
    data_filter = data_3.str.contains("MID")
    data_mid = data_f[data_filter]
    
    data_filter = data_3.str.contains("TOP")
    data_top = data_f[data_filter]
'''



'''
# find number of files in dir
file_count = count_files(csv_folder)
#count = 10 
'''

'''
# count number of raw data files  -> *could just count rows in KPI file* 
def count_files(folder):
    numfiles = len([f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f)) and f[0] != '.'])
    return numfiles 
'''

'''
# loop through data files
for i in range(1,file_count+1):
    file_temp = ("CP300_SER" + str(i) + ".csv")
    data_temp = skip_to(file_temp, "Number", csv_folder)
    
    # create filter for current row, create dataframe of current row
    cesar_rowi = cesar['Test Number'] == i
    cesar_curr = cesar[cesar_rowi]
    
    
    '''
    
    
# cwd after loop is 'files' not 'Code'
#print(os.getcwd())
    








































